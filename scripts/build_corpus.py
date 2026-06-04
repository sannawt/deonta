#!/usr/bin/env python3
"""
Build a universal runtime corpus from the compliance workbook.

Outputs in build/:
  - corpus.dl
  - predicates.json
  - citations.json
  - regulations.json
  - rules_index.json
"""

from __future__ import annotations

import argparse
import json
import os
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REPO = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = Path.home() / "Compliance calculator.xlsx"


def _json_dump(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def _trim_rows(rows: list[tuple[Any, ...]]) -> list[tuple[Any, ...]]:
    while rows and all(v is None or (isinstance(v, str) and not v.strip()) for v in rows[-1]):
        rows.pop()
    return rows


def _sheet_as_dicts(ws: Any) -> list[dict[str, Any]]:
    raw = _trim_rows(list(ws.iter_rows(values_only=True)))
    if not raw:
        return []
    headers = [str(h).strip() if h is not None else "" for h in raw[0]]
    while headers and not headers[-1]:
        headers.pop()
    out: list[dict[str, Any]] = []
    for row in raw[1:]:
        if row is None or all(c is None or (isinstance(c, str) and not str(c).strip()) for c in row):
            continue
        item: dict[str, Any] = {}
        for i, name in enumerate(headers):
            if not name:
                continue
            val = row[i] if i < len(row) else None
            if isinstance(val, float) and val == int(val):
                val = int(val)
            item[name] = val
        out.append(item)
    return out


def _slug_regulation(value: Any) -> str:
    raw = str(value or "").strip().lower()
    raw = raw.replace(" ", "_").replace("-", "_")
    return re.sub(r"[^a-z0-9_]+", "", raw)


def _replace_single_quotes(text: str) -> str:
    return re.sub(r"'([^'\n]*)'", lambda m: '"' + m.group(1).replace('"', '\\"') + '"', text)


def _split_top_level(text: str, sep: str) -> list[str]:
    parts: list[str] = []
    buf: list[str] = []
    depth = 0
    in_string = False
    quote_char = ""
    for ch in text:
        if in_string:
            buf.append(ch)
            if ch == quote_char:
                in_string = False
            continue
        if ch in ('"', "'"):
            in_string = True
            quote_char = ch
            buf.append(ch)
            continue
        if ch == "(":
            depth += 1
            buf.append(ch)
            continue
        if ch == ")":
            depth = max(0, depth - 1)
            buf.append(ch)
            continue
        if ch == sep and depth == 0:
            part = "".join(buf).strip()
            if part:
                parts.append(part)
            buf = []
            continue
        buf.append(ch)
    tail = "".join(buf).strip()
    if tail:
        parts.append(tail)
    return parts


def _normalize_rule_line(line: str) -> str:
    s = line.rstrip()
    if not s.strip():
        return ""
    if s.lstrip().startswith("%"):
        idx = s.index("%")
        return s[:idx] + "//" + s[idx + 1 :]
    s = _replace_single_quotes(s)
    s = s.replace("<=", ":-")
    s = re.sub(r"\s*&\s*", ", ", s)
    s = re.sub(r"\s*~\s*", "!", s)
    s = re.sub(r"^\+\s*", "", s.strip())
    if s and not s.endswith(".") and not s.strip().startswith("//"):
        s += "."
    return s


def _normalize_rule_cell(text: str | None) -> list[str]:
    if not text or not str(text).strip():
        return []
    lines: list[str] = []
    for raw in str(text).splitlines():
        norm = _normalize_rule_line(raw)
        if norm:
            lines.append(norm)
    return lines


def _expand_body_disjunction(rule_line: str) -> list[str]:
    if ":-" not in rule_line or "|" not in rule_line:
        return [rule_line]
    head, body = rule_line.split(":-", 1)
    body = body.rstrip(".").strip()
    alts = _split_top_level(body, "|")
    if len(alts) <= 1:
        return [rule_line]
    return [f"{head.strip()} :- {alt.strip()}." for alt in alts if alt.strip()]


def _atom_predicate(atom: str) -> str | None:
    m = re.match(r"^!?([A-Za-z_][A-Za-z0-9_]*)\s*\(", atom.strip())
    return m.group(1) if m else None


def _atom_args(atom: str) -> list[str]:
    atom = atom.strip().rstrip(".")
    start = atom.find("(")
    end = atom.rfind(")")
    if start < 0 or end < start:
        return []
    inner = atom[start + 1 : end].strip()
    if not inner:
        return []
    return [part.strip() for part in _split_top_level(inner, ",")]


def _arity_of_atom(atom: str) -> int:
    return len(_atom_args(atom))


def _variables(args: list[str]) -> set[str]:
    return {
        a
        for a in args
        if re.match(r"^[A-Z_][A-Za-z0-9_]*$", a)
    }


@dataclass
class RuleRecord:
    provision_long_id: str
    regulation: str
    scope_tag: str
    rule_text: str
    head_atom: str
    head_predicate: str
    head_arity: int
    body_atoms: list[str]
    body_predicates: list[str]
    negated_predicates: list[str]
    source_type: str


def _parse_rule_record(
    provision_long_id: str,
    regulation: str,
    scope_tag: str,
    line: str,
    source_type: str,
) -> RuleRecord | None:
    if ":-" not in line or line.strip().startswith("//"):
        return None
    head, body = line.split(":-", 1)
    head = head.strip()
    body = body.rstrip(".").strip()
    head_pred = _atom_predicate(head)
    if not head_pred:
        return None
    body_parts = [part.strip() for part in _split_top_level(body, ",") if part.strip()]
    return RuleRecord(
        provision_long_id=provision_long_id,
        regulation=regulation,
        scope_tag=scope_tag,
        rule_text=line,
        head_atom=head,
        head_predicate=head_pred,
        head_arity=_arity_of_atom(head),
        body_atoms=body_parts,
        body_predicates=[p for p in (_atom_predicate(part) for part in body_parts) if p],
        negated_predicates=[
            p for p in (_atom_predicate(part) for part in body_parts if part.startswith("!")) if p
        ],
        source_type=source_type,
    )


def _validate_range_restriction(rule_line: str) -> str | None:
    if ":-" not in rule_line or rule_line.strip().startswith("//"):
        return None
    head, body = rule_line.split(":-", 1)
    head_vars = _variables(_atom_args(head))
    body_parts = [part.strip() for part in _split_top_level(body.rstrip(".").strip(), ",") if part.strip()]
    positive_vars: set[str] = set()
    for part in body_parts:
        if part.startswith("!"):
            continue
        positive_vars.update(_variables(_atom_args(part)))
    missing = sorted(head_vars - positive_vars)
    if missing:
        return f"range restriction failed for {head.strip()}: unbound head vars {missing}"
    return None


def _tarjan(nodes: list[str], edges: dict[str, set[str]]) -> list[list[str]]:
    index = 0
    stack: list[str] = []
    on_stack: set[str] = set()
    indices: dict[str, int] = {}
    low: dict[str, int] = {}
    comps: list[list[str]] = []

    def strongconnect(v: str) -> None:
        nonlocal index
        indices[v] = index
        low[v] = index
        index += 1
        stack.append(v)
        on_stack.add(v)
        for w in edges.get(v, set()):
            if w not in indices:
                strongconnect(w)
                low[v] = min(low[v], low[w])
            elif w in on_stack:
                low[v] = min(low[v], indices[w])
        if low[v] == indices[v]:
            comp: list[str] = []
            while True:
                w = stack.pop()
                on_stack.remove(w)
                comp.append(w)
                if w == v:
                    break
            comps.append(comp)

    for node in nodes:
        if node not in indices:
            strongconnect(node)
    return comps


def _validate_stratification(rule_records: list[RuleRecord]) -> list[str]:
    nodes = sorted({r.head_predicate for r in rule_records} | {p for r in rule_records for p in r.body_predicates})
    graph: dict[str, set[str]] = defaultdict(set)
    neg_edges: set[tuple[str, str]] = set()
    for r in rule_records:
        for p in r.body_predicates:
            graph[r.head_predicate].add(p)
        for p in r.negated_predicates:
            neg_edges.add((r.head_predicate, p))
    errors: list[str] = []
    for comp in _tarjan(nodes, graph):
        comp_set = set(comp)
        for src, dst in neg_edges:
            if src in comp_set and dst in comp_set:
                errors.append(
                    f"stratification failed: negative cycle involving {src} and {dst}"
                )
    return errors


def _emit_universal_rules(regulations: list[str]) -> tuple[list[str], list[RuleRecord]]:
    lines = [
        "// ---- generated universal orchestration ----",
        'in_force(R, "current") :- entry_into_force(R, _).',
        'in_force(R, "current") :- application_date(R, "general", _).',
        "applies(S, T, R) :- in_force(R, T), regulation_territorial_link(R, A, S), regulation_material(R, S), !regulation_excluded(R, S).",
        "applies_via_actor(S, T, R, A) :- in_force(R, T), regulation_territorial_link(R, A, S), regulation_material(R, S), !regulation_excluded(R, S).",
    ]
    out_records: list[RuleRecord] = []
    for line in lines[1:]:
        rr = _parse_rule_record("__generated__", "__generated__", "ORCHESTRATION", line, "generated")
        if rr:
            out_records.append(rr)
    for reg in regulations:
        lines.append(f'regulation_material("{reg}", S) :- {reg}_material(S).')
        lines.append(
            f'regulation_territorial_link("{reg}", A, S) :- {reg}_territorial_link(A, S).'
        )
        lines.append(f'regulation_excluded("{reg}", S) :- {reg}_excluded(S).')
        for line in lines[-3:]:
            rr = _parse_rule_record("__generated__", reg, "ORCHESTRATION", line, "generated")
            if rr:
                out_records.append(rr)
    return lines, out_records


def build_corpus(xlsx: Path, out_dir: Path) -> dict[str, Any]:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    articles = _sheet_as_dicts(wb["Articles + Rules"])
    required = _sheet_as_dicts(wb["Required facts"])
    obligations: list[dict[str, Any]] = []
    citations_rows: list[dict[str, Any]] = []
    if "Obligation rules" in wb.sheetnames:
        obligations = _sheet_as_dicts(wb["Obligation rules"])
    if "Full text Datalog map" in wb.sheetnames:
        citations_rows = _sheet_as_dicts(wb["Full text Datalog map"])
    wb.close()

    if not citations_rows:
        citations_rows = [
            {
                "provision_id": row.get("provision_id"),
                "provision_long_id": row.get("provision_long_id"),
                "regulation": row.get("regulation"),
                "type": row.get("type"),
                "scope_tag": row.get("scope_tag"),
                "title": row.get("title") or row.get("provision_name"),
                "text": row.get("text"),
                "datalog_rule": row.get("datalog_rule"),
            }
            for row in articles
            if str(row.get("provision_long_id") or row.get("provision_id") or "").strip()
        ]

    regulations = sorted(
        {
            _slug_regulation(row.get("regulation"))
            for row in articles + obligations
            if str(row.get("regulation") or "").strip()
        }
    )

    if not regulations:
        raise SystemExit("No regulations discovered in workbook")

    predicate_rows: dict[str, dict[str, Any]] = {}
    for row in required:
        pred = str(row.get("predicate") or "").strip()
        if not pred:
            continue
        item = dict(row)
        item["predicate"] = pred
        try:
            item["arity"] = int(float(item.get("arity"))) if item.get("arity") is not None else 0
        except (TypeError, ValueError):
            item["arity"] = 0
        predicate_rows[pred] = item

    emitted_lines: list[str] = []
    rule_records: list[RuleRecord] = []
    arities: dict[str, int] = {p: int(r.get("arity", 0)) for p, r in predicate_rows.items()}
    errors: list[str] = []

    def ingest_rows(rows: list[dict[str, Any]], source_type: str) -> None:
        nonlocal emitted_lines, rule_records, arities, errors
        for row in rows:
            reg = _slug_regulation(row.get("regulation"))
            scope_tag = str(row.get("scope_tag") or "UNSCOPED").strip().upper()
            plid = str(row.get("provision_long_id") or row.get("provision_id") or "UNKNOWN")
            norm_lines = _normalize_rule_cell(row.get("datalog_rule"))
            for line in norm_lines:
                expanded = _expand_body_disjunction(line)
                for rule_line in expanded:
                    if rule_line.strip().startswith("//"):
                        emitted_lines.append(rule_line)
                        continue
                    head_pred = _atom_predicate(rule_line.split(":-", 1)[0])
                    if head_pred in {"applies", "applies_via_actor"}:
                        continue
                    if head_pred == "obligation" and source_type != "obligations":
                        continue
                    emitted_lines.append(rule_line)
                    if head_pred:
                        arities[head_pred] = max(arities.get(head_pred, 0), _arity_of_atom(rule_line.split(":-", 1)[0]))
                    if ":-" in rule_line:
                        rr = _parse_rule_record(plid, reg, scope_tag, rule_line, source_type)
                        if rr:
                            rule_records.append(rr)
                            body = rule_line.split(":-", 1)[1].rstrip(".").strip()
                            for atom in [part.strip() for part in _split_top_level(body, ",") if part.strip()]:
                                pred = _atom_predicate(atom)
                                if pred:
                                    arities[pred] = max(arities.get(pred, 0), _arity_of_atom(atom))
                        rr_error = _validate_range_restriction(rule_line)
                        if rr_error:
                            errors.append(rr_error)
                    else:
                        pred = _atom_predicate(rule_line)
                        if pred:
                            arities[pred] = max(arities.get(pred, 0), _arity_of_atom(rule_line))

    ingest_rows(articles, "articles")
    ingest_rows(obligations, "obligations")

    universal_lines, universal_records = _emit_universal_rules(regulations)
    emitted_lines.extend([""] + universal_lines)
    rule_records.extend(universal_records)
    for line in universal_lines:
        pred = _atom_predicate(line.split(":-", 1)[0] if ":-" in line else line)
        if pred:
            arities[pred] = max(arities.get(pred, 0), _arity_of_atom(line.split(":-", 1)[0] if ":-" in line else line))

    # Validation: predicate names and arities across rule bodies.
    for rr in rule_records:
        if rr.head_predicate not in arities:
            errors.append(f"unknown head predicate {rr.head_predicate!r}")
        for pred in rr.body_predicates:
            if pred not in arities:
                errors.append(
                    f"unknown body predicate {pred!r} in {rr.provision_long_id}"
                )

    # Shim completeness.
    head_preds = {rr.head_predicate for rr in rule_records}
    for reg in regulations:
        for shim in (f"{reg}_material", f"{reg}_territorial_link", f"{reg}_excluded"):
            if shim not in head_preds:
                errors.append(f"missing shim source predicate {shim!r} for regulation {reg}")

    errors.extend(_validate_stratification(rule_records))

    # Augment predicate catalog with inferred intensional/generated predicates.
    for name, arity in sorted(arities.items()):
        if name in predicate_rows:
            continue
        predicate_rows[name] = {
            "predicate": name,
            "arity": arity,
            "argument types": ", ".join(["symbol"] * arity),
            "kind": "generated" if name.startswith("regulation_") or name in {"applies", "applies_via_actor"} else "intensional",
            "scope dimension": "",
            "description": "Inferred from rule heads",
            "source articles": "",
            "example": "",
        }

    # Output directives for the reasoner/debugger: emit all non-extensional relations
    # so provenance/defeasibility can inspect actual derived atoms for a scenario.
    decl_lines = ["// Auto-generated universal corpus declarations."]
    input_lines: list[str] = []
    output_lines: list[str] = []
    for name in sorted(arities):
        arity = max(0, arities[name])
        args = ", ".join(f"v{i}: symbol" for i in range(arity))
        inner = f"({args})" if arity else "()"
        decl_lines.append(f".decl {name}{inner}")
        kind = str(predicate_rows.get(name, {}).get("kind") or "").strip().lower()
        if kind == "extensional":
            input_lines.append(f".input {name}")
        else:
            output_lines.append(f".output {name}")

    corpus_text = "\n".join(
        decl_lines
        + [""]
        + sorted(set(input_lines))
        + [""]
        + sorted(set(output_lines))
        + [""]
        + emitted_lines
        + [""]
    )

    citations = {
        str(row.get("provision_long_id") or row.get("provision_id") or "UNKNOWN"): {
            "provision_id": row.get("provision_id"),
            "provision_long_id": row.get("provision_long_id"),
            "regulation": _slug_regulation(row.get("regulation")),
            "type": row.get("type"),
            "scope_tag": row.get("scope_tag"),
            "title": row.get("title"),
            "text": row.get("text"),
            "datalog_rule": row.get("datalog_rule"),
        }
        for row in citations_rows
    }

    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "corpus.dl").write_text(corpus_text, encoding="utf-8")
    _json_dump(out_dir / "predicates.json", list(predicate_rows.values()))
    _json_dump(out_dir / "citations.json", citations)
    _json_dump(out_dir / "regulations.json", regulations)
    _json_dump(
        out_dir / "rules_index.json",
        [
            {
                "provision_long_id": rr.provision_long_id,
                "regulation": rr.regulation,
                "scope_tag": rr.scope_tag,
                "rule_text": rr.rule_text,
                "head_atom": rr.head_atom,
                "head_predicate": rr.head_predicate,
                "head_arity": rr.head_arity,
                "body_atoms": rr.body_atoms,
                "body_predicates": rr.body_predicates,
                "negated_predicates": rr.negated_predicates,
                "source_type": rr.source_type,
            }
            for rr in rule_records
        ],
    )

    summary = {
        "regulations": len(regulations),
        "predicates": len(predicate_rows),
        "rules": len(rule_records),
        "errors": errors,
        "xlsx": str(xlsx),
        "out_dir": str(out_dir),
    }
    return summary


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "xlsx",
        nargs="?",
        default=None,
        help="Workbook path (default: $COMPLIANCE_CORPUS_XLSX or ~/Compliance calculator.xlsx)",
    )
    parser.add_argument("-o", "--out", type=Path, default=REPO / "build")
    args = parser.parse_args()

    xlsx = (
        Path(args.xlsx).expanduser().resolve()
        if args.xlsx
        else Path(
            os.environ.get("COMPLIANCE_CORPUS_XLSX", str(DEFAULT_XLSX))
        ).expanduser().resolve()
    )
    if not xlsx.is_file():
        raise SystemExit(f"Workbook not found: {xlsx}")

    summary = build_corpus(xlsx, args.out)
    print(
        f"{summary['regulations']} regulations · {summary['predicates']} predicates · "
        f"{summary['rules']} rules · {len(summary['errors'])} errors"
    )
    if summary["errors"]:
        for err in summary["errors"]:
            print(f"ERROR: {err}")
        raise SystemExit(1)


if __name__ == "__main__":
    main()
